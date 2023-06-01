'''
自动把成绩打印到试卷上
试卷放bak，已设好区域坐标，指定有成绩的excel，二卷成绩在表名为erjuan,考号在第一列，总分在最后一列
'''
import json
import cv2,sys
import openpyxl
from pathlib import Path
imagePath = Path(sys.argv[0]).parent
imagepic = imagePath.joinpath('bak')#1.放试卷处
quyulistfenx=[]
quyulistfeny=[]


with open('setting.json', 'r', encoding='utf-8') as json_file:
    result = json.load(json_file)

def fentopic():
    try:
        chengjiexcel = openpyxl.load_workbook(result['self.openexcelname'])  # 打开excel
    except:
        print('excel文件损坏！')
    if 'quyufs' in chengjiexcel.sheetnames:
        sh5 = chengjiexcel['quyufs']  # 打开表单三
        shcs=chengjiexcel['cs']
    else:
        print('没有quyufs表，选采集打分区域！')
    quyulistfenx.clear()
    quyulistfeny.clear()
    for i in range(sh5.max_column):  # 共105题，从2行到107行
        quyulistfenx.append(sh5.cell(row=1, column=i + 1).value)
        quyulistfeny.append(sh5.cell(row=2, column=i + 1).value)
    print(quyulistfenx)
    for fielname in [x for x in imagepic.iterdir()]:
        print(fielname.name)
        img = cv2.imread(str(fielname))
        if shcs.cell(row=12,column=2).value != None:
            cut_img=img[quyulistfeny[0]:quyulistfeny[1], quyulistfenx[0]:quyulistfenx[1]]#3条形码y1:y2,x1:x2
            if len(img.shape) == 2:
                img = cv2.cvtColor(img, cv2.COLOR_GRAY2BGR)
            if len(cut_img.shape) == 3:  # 如果是彩色的就化为灰度图
                gray_image = cv2.cvtColor(cut_img, cv2.COLOR_BGR2GRAY)
            else:
                gray_image = cut_img  # 否则不转化
            import pic2tiaoxingma
            kauanohaoch = pic2tiaoxingma.gettiaoxingma(gray_image)
            print(kauanohaoch)
            sh3 = chengjiexcel['erjuan']#2.指定名表
            for i in range(3, sh3.max_row + 1):  # 从第3行开始，到第3行结束
                xuehao = int(sh3.cell(row=i, column=1).value)#2.指定考号列为1
                if str(kauanohaoch) in str(xuehao) :  # 两号相同才画错题，写存文件
                    #4成绩列数，5对应写到的坐标。                          总分所在列数最后一列，总分区域坐标也在最后
                    cv2.putText(img, '{} '.format(sh3.cell(row=i, column=sh3.max_column).value),
                                (quyulistfenx[-1], quyulistfeny[-1]), cv2.FONT_HERSHEY_TRIPLEX, 1,
                                (0, 0, 255), 2)
                    for j in range(2,sh3.max_column):#5打印每列即每一题的小分
                        if sh3.cell(row=i, column=j).value !=None:
                            cv2.putText(img, '{} '.format(sh3.cell(row=i, column=j).value),
                                    (quyulistfenx[j], quyulistfeny[j]), cv2.FONT_HERSHEY_TRIPLEX, 1,
                                    (0, 0, 255), 2)

                    cv2.imwrite(str(fielname), img)
        else:#如果没有考号
            sh3 = chengjiexcel['erjuan']  # 2.指定名表
            for i in range(3, sh3.max_row + 1):  # 从第2行开始，到第3行结束
                xuehao = sh3.cell(row=i, column=1).value  # 2.指定考号列为1
                if fielname.name == str(xuehao):  # 两号相同才画错题，写存文件
                    # 4成绩列数，5对应写到的坐标。                          总分所在列数
                    cv2.putText(img, '{} '.format(sh3.cell(row=i, column=sh3.max_column).value),
                                (quyulistfenx[0], quyulistfeny[0]), cv2.FONT_HERSHEY_TRIPLEX, 1,
                                (0, 0, 255), 2)
                    for j in range(2, sh3.max_column):  # 5打印每列即每一题的小分
                        if sh3.cell(row=i, column=j).value != None:
                            cv2.putText(img, '{} '.format(sh3.cell(row=i, column=j).value),
                                        (quyulistfenx[j - 1], quyulistfeny[j - 1]), cv2.FONT_HERSHEY_TRIPLEX, 1,
                                        (0, 0, 255), 2)

                    cv2.imwrite(str(fielname), img)
if __name__ == "__main__":
    fentopic()
    print('分数全部打印在试卷上了！')