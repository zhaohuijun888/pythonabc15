#生成错题集doc后，自动发送邮件并删除生成的doc.被ctmain2email3调用。
import smtplib
import time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
import sys
import openpyxl
imagePath = Path(sys.argv[0]).parent

def fsem(zdname,wjdzhi):
    lctime = time.localtime()

    ymdtime = time.strftime("%Y-%m-%d",lctime)
    wb = openpyxl.load_workbook(imagePath.joinpath('moban.xlsx'))#获取名字，email，二卷
    ws = wb['yijuan']
    emaildz=wb['email']
    rows = ws.max_row
    # print(rows)
    for i in range(3,rows + 1):#从第2行开始，到第3行结束
        xsname = ws.cell(row=i, column=2).value
        email = emaildz.cell(row=i, column=2).value
        wjdzhi=ws.cell(row=i, column=ws.max_column).value
        if wjdzhi != None and email != None and zdname==xsname:
            fujian1 = wjdzhi
            fujian2=imagePath.joinpath('wuli.docx')
            # 1.创建一个带附件的实例
            msg = MIMEMultipart()
            #2. 加邮件头
            msg['From'] = '87336683@qq.com'
            msg['To'] = email
            msg['Subject'] = "高三（9）班物理{}试卷".format(ymdtime) + xsname
            # 3.构造附件1
            att1 = MIMEText(open(fujian2, 'rb').read(), 'base64', 'utf-8')
            att1["Content-Type"] = 'application/octet-stream'
            att1["Content-Disposition"] = 'attachment; filename="wuliti.docx"'
            msg.attach(att1)
            # 3构造附件2（附件为JPG格式的图片）
            att2 = MIMEText(open(fujian1, 'rb').read(), 'base64', 'utf-8')
            att2["Content-Type"] = 'application/octet-stream'
            att2["Content-Disposition"] = 'attachment; filename="wulisj.jpg"'
            msg.attach(att2)
            # 4.发送邮件
            server = smtplib.SMTP_SSL('smtp.qq.com',465)
            # print('1')
            try:
                # print('2')
                server.login('87336683@qq.com','fkniwxtdvecnbhbg')
                # print('3')
                server.send_message(msg)#发邮箱、接邮箱、纯文本内容
                print('已发送{}'.format(xsname))
                server.quit()
                time.sleep(1)
            except Exception:
                print('发送{}失败！！！！！！！！！！'.format(xsname))
                pass


if __name__ == "__main__":
    fsem()
