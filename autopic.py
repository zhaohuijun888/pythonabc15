# 28行设置开始行数，55，56设置采集试题还是答案
import time
import winsound
import keyboard
import openpyxl
from PIL import ImageGrab
import ctypes
from pathlib import Path
import sys

from win32com.client import Dispatch

speaker = Dispatch('SAPI.SpVoice')
imagePath = Path(sys.argv[0]).parent
lctime = time.localtime()
tishistr = '按n开始截图，按q退出'
namestr = ''
tihao = ''
cuotilist = []
jishuqi = 0
hangkaiguan = 2
shitidaankaiguan = 'st'
hangkaiguan, shitidaankaiguan = input('请输入开始采集行数和采集类型如20 dn、15 st:').split()
hangkaiguan = int(hangkaiguan)


def listst():
    try:
        chengjiexcel = openpyxl.load_workbook('ct.xlsx')  # 打开excel
        print('文件已打开')
        sh5 = chengjiexcel['ly']  # 打开表单三
        print('表已打开')
        rows = sh5.max_row
        for i in range(hangkaiguan, rows + 1):  # 从1题开始
            # cuotilist.append(sh5.cell(row=i, column=5).value)
            cuotilist.append(sh5.cell(row=i, column=1).value)
        if len(cuotilist) > 0:
            print('请准备{}截取'.format(cuotilist[0]))
            speaker.Speak('请准备{}截取'.format(
                cuotilist[0].replace('zx', '摘星').replace('gg', '巩固').replace('yk', '月考').replace('mn', '模拟')[:-4]))
        else:
            pass
    except:
        print('ct.xlsx不存在！')


def capture():
    try:
        dll = ctypes.cdll.LoadLibrary('PrScrn.dll')
    except Exception:
        print("Dll load error!")
        return
    else:
        try:
            dll.PrScrn(0)
        except Exception:
            print("Sth wrong in capture!")
            return


def test_n():
    global namestr, tihao, jishuqi, cuotilist
    if shitidaankaiguan == 'st':
        liststr = str(cuotilist[jishuqi])  # 读取错题题号例表
    else:
        liststr = str(cuotilist[jishuqi]).replace('.', 'a.')  # 读取错题题号答案例表
    capture()
    time.sleep(0.1)
    image = ImageGrab.grabclipboard()  # 获取剪切板的图片

    print(liststr)

    image.save('./img/{}'.format(liststr))
    print(namestr, tihao)

    print('jsq', jishuqi, )
    if jishuqi < len(cuotilist) - 2:
        print('请准备{}截取'.format(cuotilist[jishuqi + 1]))
        speaker.Speak('请准备{}截取'.format(
            cuotilist[jishuqi + 1].replace('zx', '摘星').replace('zk', '周考').replace('zt', '专题').replace('gg',
                                                                                                       '巩固').replace(
                'yk', '月考').replace('mn', '模拟')[:-4]))
        jishuqi = jishuqi + 1
    else:
        print('请准备{}截取,这是最后一题了！！！！！'.format(cuotilist[jishuqi + 1]))
        speaker.Speak('请准备{}截取,这是最后一题了！！！！！'.format(
            cuotilist[jishuqi + 1].replace('zx', '摘星').replace('zk', '周考').replace('zt', '专题').replace('gg',
                                                                                                       '巩固').replace(
                'yk', '月考').replace('mn', '模拟')[:-4]))


def main():
    if lctime.tm_year <= 2023:
        listst()
        print(tishistr)
        keyboard.add_hotkey('n', test_n)  # 自动保试卷截图图片，需要输入名字
        # keyboard.add_hotkey('q', test_q)
        keyboard.wait('q')  # 按q键退出
    else:
        pass


if __name__ == '__main__':
    main()
