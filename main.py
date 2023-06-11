import json
import time
from itertools import chain

import imutils
import numpy as np

import pyzbar.pyzbar as pyzbar
import cv2
import openpyxl
from PIL import Image, ImageDraw
from PyQt5 import QtWidgets
from PyQt5.QtCore import pyqtSlot
from PyQt5.QtGui import QImage, QPixmap
from PyQt5.QtWidgets import QMainWindow, QDesktopWidget, QFileDialog
from openpyxl.styles import Font
from need.untitled import Ui_MainWindow
import logging
import glob
import os
import shutil
from pathlib import Path
from pystrich.code128 import Code128Encoder

imagePath = Path.cwd()


class MainWindow(QMainWindow, Ui_MainWindow):
    def __init__(self, parent=None):
        super(MainWindow, self).__init__(parent)
        self.setupUi(self)
        self.screen = QDesktopWidget().screenGeometry()  # 获取屏幕尺寸
        self.move(0, 0)  # 窗体放中心
        self.resize(self.screen.width(), self.screen.height() - 105)
        # 去掉下面注释，不显示logging,加上注释显示
        # logging.disable(logging.CRITICAL)#正常不显示错误信息，加上注释显示
        logging.basicConfig(level=logging.DEBUG)
        self.list1 = set()  # 关注名单集合
        self.duoxuanset = set()  # 单选题，判为多选文件名集合
        # 阈值初使化
        self.yuzhi = 210
        # 打开excel文件名变量初化
        self.openexcelname = ''
        # 存储答案列表初始化
        self.answer = [''] * 21
        # 选项坐标初使化
        self.quyulist = []
        # 考号坐标初使化
        self.quyulist1 = []
        # 初始化拼接开关
        self.pinjiekaiguan = 0
        self.jianquekg = True  # 剪切试卷开关
        # 试卷列表清空
        self.filelist = []
        # 试卷文件列表指针
        self.filenumjsq = 0
        # 批阅完成试卷计数
        self.youxiaonum = 0
        # 选项中有空白试卷文件名
        self.kongxuanxiansum = ''
        # 有效试卷总分
        self.zongfenbj = 0
        # 涂黑面积所占百分比例设置
        self.mianjibaifenbi = 30
        # 识别成功的考号
        self.kauanohaoch = []
        # 识别出的考号单个字符
        self.kaohao = []
        # 答对题目数量
        self.right = 0
        # 半对题目数量
        self.bandui = 0
        # 空白题目数量
        self.kong = 0
        # 条形码最小值和最大值
        self.zimu = [0, 0]
        # 挑选试题开关
        self.tiaoxuanshijuankaiguan = False
        self.wenjianjia = ''
        self.tiaoxingmacw = ''  # 条形码错误提示。
        # 允许自动调整区域
        self.dingweikaiguan = True
        try:
            # 加载最近一次设置的文件和文件夹变量
            self.moernset()
            # 加载最近一次文件内的试卷参数
            self.quyulistsc()
        # 如果出错，加载默认模板
        except:
            self.openexcelname = ''

            self.openexcelname = 'moban.xlsx'
            self.wenjianjia = 'ls'
            new_dict = {"self.openexcelname": "moban.xlsx", "self.wenjianjia": "ls"}  # 新建数据
            with open("setting.json", 'w') as f:  # 新建JSON文件
                json.dump(new_dict, f)  # 写入数据
            print('setting.json文件出错，已重设且默认文件为moban.xlsx，文件夹为ls')
            self.quyulistsc()

        self.helpstr = """使 用 说 明
1.修改moban.xlsx内学生考号、姓名、试题号、试题答案，按cs表内要求设置试卷信息或用步骤2的方法或鼠标设置。
2.在第1步设置过cs试卷信息的可以跳过第2步。
  1）单击【选择excel】按钮选择excel.
  2）单击【重新输入】按钮，输入选择题数、选择总分、每题分数、部分分数及每题正确答案，输入完毕按【确认答案】.
  3）单击【选择文件夹】按钮，选择试卷所在文件夹。
  4）坐标采集方法在采集程序内。  
3.单击【选择文件夹】按钮，选择试卷所在文件夹。
4.单击【调整区域】按钮预览区域.第1个文本输入如10 10单击【调整区域】调整选项区域，输入-2 3 1只调整所有区域。
5.单击【开始阅卷】按钮，开始阅卷。第一个文本输入框输入数字1是正常阅卷，2是查看二值图，3是问题试卷强制通过。
6. 试卷阅完后，第一个单行文本输入框内输入学生姓名，单击【查询学生】按钮进行查看。
7. 单击【导出excel】按钮统计分析学生成绩。
8.第二个文本输入框输入['15','20','25','30','45','60']中的数字单击阈值设置，可调整识别比率。正常30，单选判为多选可调大，空白选项多时调小。输入其它数学，可调整阈值。
9.阈值输入框依次输入ej、xt、cj、ct、yj、bz,txsj,txm单击【阈值设置】按钮分别启动批改二卷、打小题分、采集试题库、
生成单个学生错题库、自动发送所有学生发试题和错题本邮件、打开帮助、挑选试卷（放在PIC目录下,单击阈值设置,开始阅卷按纽,完成后在ls目录下）,修补条形码（第一个输入框输入条形码后两位用空格分隔，第二个输入框输入txm后单击阈值设置按钮再单击开始阅卷）等功能。
10.源码：https://github.com/zhaohuijun888/pythonabc15
更多功能、操作演示请观看视频演示（B站：赵会钧）。

        """
        # 标签显示帮助
        self.label_shijuan.setText(self.helpstr)
        # 允许换行显示
        self.label_shijuan.setWordWrap(True)
        # self.label_shijuan.setAlignment(Qt.AlignLeft,Qt.AlignTop,Qt.)
        # 设置字体大小
        if self.screen.width() >= 2160:
            self.label_shijuan.setStyleSheet("font-size:35px;")
        else:
            self.label_shijuan.setStyleSheet("font-size:15px;")
        logging.debug('准备遍历试卷')
        # 试卷列表清空
        self.filelist = []
        self.bianlipic()
        logging.debug('遍历试卷完成')
        # 打印常规信息，提供出错诊断
        if len(self.filelist) > 0:
            print('试卷格式为：' + str(self.filelist[self.filenumjsq])[-3:])
            if str(self.filelist[self.filenumjsq])[-3:] != 'jpg':
                print('图片格式不是jpg!')
                # sys.exit(0)
        print('试卷份数为：{}'.format(len(self.filelist)))
        print('excel文件是：', self.openexcelname)
        print('试题个数与答案个数为：', self.tishu, len(self.answer))
        print('试卷文件夹为：', self.wenjianjia)

    def shengchengtiaoxingma(self):
        logging.debug(self.lineEdit_chaxunxuesheng.text())
        tiaoxingmalist = self.lineEdit_chaxunxuesheng.text().split()
        logging.debug('进入制作条形码程序')
        self.bianlipic()
        for i in range(len(tiaoxingmalist)):
            logging.debug(self.zimu[0])
            logging.debug(str(self.zimu[0])[:-2])
            logging.debug(type(tiaoxingmalist[i]))
            if len(tiaoxingmalist[i]) == 1:
                tiaoxingmalist[i] = str(self.zimu[0])[:-1] + tiaoxingmalist[i]
            else:
                tiaoxingmalist[i] = str(self.zimu[0])[:-2] + tiaoxingmalist[i]
            logging.debug(tiaoxingmalist[i])
            encoder = Code128Encoder(tiaoxingmalist[i],
                                     options={"ttf_font": "C:/Windows/Fonts/SimHei.ttf", "ttf_fontsize": 22,
                                              "bottom_border": 5, "height": 150, "label_border": 1})  # 生成
            encoder.save(f"{tiaoxingmalist[i]}.png", bar_width=3)
            logging.debug('生成条形码完成')
            im1 = open(str(self.filelist[i]), 'rb')
            im2 = open(f"{tiaoxingmalist[i]}.png", 'rb')
            Im1 = Image.open(im1)
            Im2 = Image.open(im2)
            logging.debug(self.quyulist1)
            Im2 = Im2.resize((self.quyulist1[2] - self.quyulist1[0], self.quyulist1[3] - self.quyulist1[1]))
            Im1.paste(Im2, (self.quyulist1[0], self.quyulist1[1]))
            logging.debug(self.filelist[i])
            Im1.save(self.filelist[i])
            print(f'条形码{tiaoxingmalist[i]}已修复。')
            # Im1.save(Path(self.filelist[i]))
            # Im1.save('pic/text.jpg')
            im1.close()
            im2.close()
            os.remove(f"{tiaoxingmalist[i]}.png")

    # 写入JSON
    def xiejson(self, jian, zhi, *args, **kwargs):
        logging.debug("键为：{}".format(jian))
        logging.debug("值为：{}".format(zhi))
        # print(type(jian))
        # 读取数据
        with open("setting.json", 'r', encoding='utf-8') as load_f:
            logging.debug("打开JSON成功")
            load_dict = json.load(load_f)
            logging.debug("加载JSON数据成功")
        # 修改数据
        load_dict[jian] = zhi
        logging.debug("数据修改成功")
        # 保存数据
        with open("setting.json", "w") as f:
            json.dump(load_dict, f)
            logging.debug("数据保存成功")

    # 读取JSON默认设置
    def moernset(self):
        with open('setting.json', 'r', encoding='utf-8') as json_file:
            result = json.load(json_file)
            logging.debug("打开JSON成功")
            self.openexcelname = result['self.openexcelname']
            logging.debug(self.openexcelname)
            if Path(self.openexcelname).is_file():
                pass
            else:
                print('文件{}不存在'.format(self.openexcelname))
                self.openexcelname = ''

            logging.debug(self.wenjianjia)
            # 获取项目目录
            logging.debug(os.getcwd() + '\\' + result['self.wenjianjia'])
            self.wenjianjia = os.getcwd() + '\\' + result['self.wenjianjia']
            if Path(self.wenjianjia).is_dir():
                pass
            else:
                print('文件夹{}不存在'.format(self.wenjianjia))

    # 设置阈值
    @pyqtSlot()
    def on_pushButton_yuzhi_clicked(self):
        yuzhistr = self.lineEdit_yuzhi.text()
        xsxm = self.lineEdit_chaxunxuesheng.text()
        logging.debug('输入阈值文本框内容为：{}'.format(yuzhistr))
        # 试卷打印小题分
        if yuzhistr == 'xt':
            import fentopic
            fentopic.fentopic()
            print('分数已全部打印到试卷上')
        # 打开II卷批改程序
        elif yuzhistr == 'ej':
            os.system('ej.exe')
        #
        # 采集错题pic
        elif yuzhistr == 'cj':
            import autopic
            autopic.main()
        # 导出word错题集
        elif yuzhistr == 'ct':
            import ctmain2
            ctmain2.main()
            print('已生成错题word。')
        # 导出word错题集并发邮件给学生
        elif yuzhistr == 'yj':
            import ctmain2email3
            ctmain2email3.main()
            print('邮件已全部发送！')
        # 帮助信息
        elif yuzhistr == 'bz':
            self.label_shijuan.setText(self.helpstr)
        elif yuzhistr == 'txm':
            self.shengchengtiaoxingma()
        elif yuzhistr == 'txsj':
            self.tiaoxuanshijuankaiguan = True
            print(self.tiaoxuanshijuankaiguan)
        elif yuzhistr in ['15', '30', '45', '60']:
            self.mianjibaifenbi = int(self.lineEdit_yuzhi.text())
            print(f'比率已设置为{self.mianjibaifenbi}')
        # 设置阈值
        else:
            self.yuzhi = int(self.lineEdit_yuzhi.text())
            print("阈值已设为{}白多调小，白少调大。".format(yuzhistr))

    # 文件选择按钮
    @pyqtSlot()
    def on_pushButton_xuanzewenjian_clicked(self):
        self.openexcelname = QFileDialog.getOpenFileName(self, '选择excel文件', "./", "excel (*.xlsx)")[0].split('/')[
            -1]
        print(self.openexcelname)
        if self.openexcelname == '':
            self.openexcelname = str(Path.cwd().joinpath('moban.xlsx'))
            print('没有选择文件，已设置为默认模板moban.xlsx')
        self.xiejson(jian='self.openexcelname', zhi=self.openexcelname)
        logging.debug('excel文件为：{}'.format(self.openexcelname))
        self.quyulistsc()

    # 重新输入
    @pyqtSlot()
    def on_pushButton_chongxinshuru_clicked(self):
        # 如果第一个文本输入框为数字1，清除所有答案
        if self.lineEdit_chaxunxuesheng.text() == '1':
            for i in range(1, 22):
                exec('self.lineEdit_0{:02}.clear()'.format(i))

        # 否则，不清除，只是能修改，原答案还在
        for i in range(1, 22):
            exec('self.lineEdit_0{:02}.setReadOnly(False)'.format(i))

        self.lineEdit_tishu.setReadOnly(False)
        self.lineEdit_zongfenshu.setReadOnly(False)
        self.lineEdit_meitifenshu.setReadOnly(False)
        self.lineEdit_bufendefen.setReadOnly(False)
        # 设置灰色样式
        for i in range(1, 22):
            exec('self.lineEdit_0{:02}.setStyleSheet("background-color:rgb(255, 255, 255)")'.format(i))

        self.lineEdit_tishu.setStyleSheet("background-color:rgb(255, 255, 255)")
        self.lineEdit_zongfenshu.setStyleSheet("background-color:rgb(255, 255, 255)")
        self.lineEdit_meitifenshu.setStyleSheet("background-color:rgb(255, 255, 255)")
        self.lineEdit_bufendefen.setStyleSheet("background-color:rgb(255, 255, 255)")

    # 答案确认按钮
    @pyqtSlot()
    def on_pushButton_daanqueren_clicked(self):
        # 清除原有答案列表
        self.answer.clear()
        # 打开excel及两个有答案的表
        try:
            self.chengjiexcel = openpyxl.load_workbook(self.openexcelname)
            self.shtongjish = self.chengjiexcel['tongjish']
            self.shyijuan = self.chengjiexcel['yijuan']
        except:
            print('存储答案时excel文件或表损坏！')
        # 清除excel两表中原有答案
        self.shyijuan.delete_cols(3, 20)
        self.shtongjish.delete_cols(1, 20)
        logging.debug('存储时的清除excel成功')
        print('存储时的清除excel成功')
        # 打开参数表cs，如果不存在则创建
        if 'cs' in self.chengjiexcel.sheetnames:
            self.shcs = self.chengjiexcel['cs']
            logging.debug('存储时的cs已打开')
        else:
            self.shcs = self.chengjiexcel.create_sheet('cs')
            logging.debug('存储时的cs已打开')

        # 试卷参数赋值变量
        self.tishu = int(self.lineEdit_tishu.text())  # 设置题目数量
        self.meitifenshu = int(self.lineEdit_meitifenshu.text())
        self.bufendefen = int(self.lineEdit_bufendefen.text())
        self.yuzhi = int(self.lineEdit_yuzhi.text())
        if self.lineEdit_chaxunxuesheng.text().isdigit():
            self.kk = (int(self.lineEdit_chaxunxuesheng.text()) - 1) * 16
        else:
            print('错误！！！开始涂卡位置不是数字，请重输入！！！或已默认设为1.')
            self.kk = 1
        # 试卷参数存入excel
        self.shcs.cell(row=1, column=2).value = self.lineEdit_tishu.text()
        self.shcs.cell(row=2, column=2).value = self.lineEdit_zongfenshu.text()
        self.shcs.cell(row=3, column=2).value = self.lineEdit_meitifenshu.text()
        self.shcs.cell(row=4, column=2).value = self.lineEdit_bufendefen.text()
        self.shcs.cell(row=1, column=1).value = '题数'
        self.shcs.cell(row=2, column=1).value = '总分'
        self.shcs.cell(row=3, column=1).value = '每题分'
        self.shcs.cell(row=4, column=1).value = '部分分'
        self.shcs.cell(row=5, column=1).value = '阈值'
        self.shcs.cell(row=6, column=1).value = '比率'
        self.shcs.cell(row=7, column=1).value = '开始卡号'
        # 答案存入答案列表self.answer
        for i in range(1, 22):
            exec('self.answer.append(self.lineEdit_0{:02}.text())'.format(i))

        # 使答案文本输入框只读状态，不能操作
        for i in range(1, 22):
            exec('self.lineEdit_0{:02}.setStyleSheet("background-color:rgb(230, 230, 230)")'.format(i))
            exec('self.lineEdit_0{:02}.setReadOnly(True)'.format(i))

        self.lineEdit_tishu.setStyleSheet("background-color:rgb(230, 230, 230)")
        self.lineEdit_tishu.setReadOnly(True)
        self.lineEdit_zongfenshu.setStyleSheet("background-color:rgb(230, 230, 230)")
        self.lineEdit_zongfenshu.setReadOnly(True)
        self.lineEdit_meitifenshu.setStyleSheet("background-color:rgb(230, 230, 230)")
        self.lineEdit_meitifenshu.setReadOnly(True)
        self.lineEdit_bufendefen.setStyleSheet("background-color:rgb(230, 230, 230)")
        self.lineEdit_bufendefen.setReadOnly(True)
        # 切片题目数量
        self.answer = self.answer[: self.tishu]
        # 答案存入excel两个表内
        for i in range(1, 22):
            exec('self.shyijuan.cell(row=2, column={}).value = self.lineEdit_0{:02}.text()'.format(i + 3, i))

        for i in range(1, 22):
            exec('self.shtongjish.cell(row=3, column={}).value = self.lineEdit_0{:02}.text()'.format(i, i))

        # 设置试卷判卷范围，一般是全判，这块功能可以去掉
        if self.lineEdit_chaxunxuesheng.text().isdigit():
            self.kk = 16 * (int(self.lineEdit_chaxunxuesheng.text()) - 1)
        else:
            print('请输入开始题号。')
        for i in range(len(self.answer)):
            # 两个表中答案加上试题编号
            self.shtongjish.cell(row=2, column=i + 1).value = '{}'.format(i + 1)
            self.shyijuan.cell(row=1, column=i + 4).value = '{}'.format(i + 1)
        # 加上I卷两个字和I卷总分
        self.shyijuan.cell(row=1, column=3).value = 'I卷'
        self.shyijuan.cell(row=2, column=3).value = self.tishu * self.meitifenshu
        # 保存excel
        try:
            self.chengjiexcel.save(self.openexcelname)
            self.chengjiexcel.close()
        except:
            print('保存出错，先关闭excel再重试！')

    # 选择文件夹按钮
    @pyqtSlot()
    def on_pushButton_xuanzeshijuan_clicked(self):
        # 选择试卷所在文件夹
        logging.debug("准备选择文件夹")
        self.wenjianjia = QFileDialog.getExistingDirectory(self, '选择学生试卷文件夹', './')
        logging.debug("已做选择")
        if self.wenjianjia == '':
            self.wenjianjia = str(Path.cwd() / 'ls')
            # self.xiejson(jian='self.wenjianjia', zhi=self.wenjianjia)
        self.xiejson(jian='self.wenjianjia', zhi=self.wenjianjia.split('/')[-1])
        logging.debug("写JOSN成功")

        # 删除bak目录下所有指定文件
        for img in glob.glob("bak/*.jpg"):
            os.remove(img)
        logging.debug("清理bak目录成功")
        # 删除pic子目录下所有指定文件
        for img in glob.glob("pic/*.jpg"):
            os.remove(img)
        logging.debug("清理pic目录成功")
        # 复制所有指定文件夹下试卷到pic目录下
        for fielname in [x for x in Path.cwd().joinpath(self.wenjianjia).iterdir()]:
            logging.debug(fielname)
            # 如果不是jpg格式，则转化为jpg
            if 'jpg' not in str(fielname):
                fname0 = fielname.name  # 取出一个的名字
                logging.debug(fname0)
                listname1 = fname0.split(".")
                filename2 = listname1[0]
                logging.debug(filename2)
                my_file = imagePath.joinpath('pic').joinpath(filename2 + '.jpg')
                logging.debug(str(my_file))
                fielname = str(fielname)
                if self.pinjiekaiguan == 1:
                    baocwenjian = ''.join(filter(lambda c: ord(c) < 256, str(my_file))).replace('pic', 'bak')

                else:
                    baocwenjian = ''.join(filter(lambda c: ord(c) < 256, str(my_file)))
                logging.debug(baocwenjian)

                logging.debug(fielname)
                os.system(f'ffmpeg -i {fielname}  {baocwenjian} -loglevel quiet')
                print(f'已转化{baocwenjian}')

            else:  # 如果是jpg直接拷到pic目录下
                logging.debug('试卷格式是jpg,拼接开关的值为：{}'.format(self.pinjiekaiguan))
                if self.pinjiekaiguan == 1:
                    logging.debug('试卷是{}'.format(fielname))
                    logging.debug('复制目的地为：{}'.format(Path.cwd().joinpath('bak').joinpath(
                        ''.join(filter(lambda c: ord(c) < 256, fielname.name)))))
                    shutil.copy(fielname, Path.cwd().joinpath('bak').joinpath(
                        ''.join(filter(lambda c: ord(c) < 256, fielname.name))))
                else:
                    shutil.copy(fielname, Path.cwd().joinpath('pic').joinpath(
                        ''.join(filter(lambda c: ord(c) < 256, fielname.name))))

        if self.pinjiekaiguan == 1:  # 左右拼接试卷
            logging.debug("准备拼接")
            listdir = imagePath.joinpath('bak')
            logging.debug('bak目录是{}'.format(listdir))
            piclsit = []
            # 遍历文件夹中每个文件名形成列表
            for fname in [x for x in listdir.iterdir()]:
                piclsit.append(fname)
            logging.debug('bak下文件列表是{}'.format(piclsit))
            try:  # pil的image有打开，没有关闭，所以改用open打开成二进制，再用pil的image的打开，然后再用close关闭open打开的图片，才能删除图片
                logging.debug('将打开{}'.format(str(piclsit[0])))
                im0 = open(str(piclsit[0]), 'rb')
                logging.debug('试卷{}打开正常'.format(str(piclsit[0])))
                Im0 = Image.open(im0)
            except:
                print('bak下没有图片,程序退出。')
                sys.exit()
            for i in range(0, len(piclsit), 2):
                im1 = open(str(piclsit[i]), 'rb')
                im2 = open(str(piclsit[i + 1]), 'rb')
                Im1 = Image.open(im1)
                Im2 = Image.open(im2)
                if Im0.size[0] > Im0.size[1]:  # 如果宽>高，剪切一部分拼接（理综）
                    logging.debug("准备拼接理综")
                    result = Image.new(Im1.mode, (Im1.size[0], Im1.size[1]))  # 生成空白
                    result.paste(Im1, (0, 0))  # 粘第一张
                    logging.debug("粘第一张成功")
                    logging.debug('空间高度{}'.format(Im1.size[1] - self.pj1y))
                    logging.debug('对象高度{}'.format(self.pj3y - self.pj2y))
                    logging.debug('对象宽度{}'.format(self.pj3x - self.pj2x))
                    # 如果剩余空间大于粘入对象，直接粘贴
                    if (Im1.size[1] - self.pj1y) >= (self.pj3y - self.pj2y):
                        result.paste(Im2.crop((self.pj2x, self.pj2y, self.pj3x, self.pj3y)), (self.pj1x, self.pj1y))
                        draw = ImageDraw.Draw(result)
                        draw.rectangle((self.pj1x, self.pj1y + self.pj3y - self.pj2y,
                                        self.pj1x + self.pj3x - self.pj2x, Im0.size[1] - 100), fill='white')
                    # 否则缩小后粘贴
                    else:
                        result.paste(Im2.crop((self.pj2x, self.pj2y, self.pj3x, self.pj3y)).resize(
                            (self.pj3x - self.pj2x, Im1.size[1] - self.pj1y)), (self.pj1x, self.pj1y))
                    logging.debug("粘第2张成功")


                else:  # 物理单科，直接拼接
                    result = Image.new(Im1.mode, (Im1.size[0] * 2, Im1.size[1]))
                    result.paste(Im1, (0, 0))
                    result.paste(Im2, (Im1.size[0], 0))
                    logging.debug('直接拼接正确')
                print('正在拼接{}'.format('{:02}.jpg'.format(int(i / 2) + 1)))
                result.save(imagePath.joinpath('pic').joinpath('{:02}.jpg'.format(int(i / 2) + 1)))
                print('正在拼接{}'.format('{:02}.jpg'.format(int(i / 2) + 1)))
                logging.debug('存新图正确')
                im0.close()
                im1.close()
                im2.close()
                logging.debug('已关闭打开的两图')
        # 删除临时文件
        for img in glob.glob("bak/*.jpg"):
            os.remove(img)
            logging.debug('删除bak图片正确')
        logging.debug('准备遍历试卷')
        # 试卷列表清空
        self.filelist = []
        self.bianlipic()
        logging.debug('遍历试卷完成')
        print('试卷准备完毕共{}份试卷。'.format(len(self.filelist)))

    # 坐标自动调整
    def quyuauto(self):
        logging.debug('进入区域自动调整')
        logging.debug(f'{self.filelist[self.filenumjsq]}进入区域自动调整')
        logging.debug(f"{self.dw1x}, {self.dw1y}, {self.dw2x}, {self.dw2y}")
        try:
            autopic = self.kaiyunsuan[self.dw1x:self.dw2x, self.dw1y:self.dw2y]
        except:
            print("抱歉，试卷格式不支持！用格式化工厂去除文件名的中文、转化为jpg后再试。")
            sys.exit(app.exec_())
        logging.debug('已选定位块区域')
        edged = cv2.Canny(autopic, 128, 200)  # 边缘检测
        logging.debug('边缘检测已成功')
        cnts = cv2.findContours(edged, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)  # 查找轮廓
        logging.debug('查找轮廓已成功')
        cnts = cnts[1] if imutils.is_cv2() else cnts[0]  # # 用以区分OpenCV2.4和OpenCV3
        if len(cnts) > 0:  # 确保至少有一个轮廓被找到
            logging.debug(f'轮廓个数为{len(cnts)}')
            cnts = sorted(cnts, key=cv2.contourArea, reverse=True)  # 将轮廓按大小降序排序
            for c in cnts:  # 对排序后的轮廓循环处理
                x, y, w, h = cv2.boundingRect(c)
                if w > 25 and h > 25:
                    x = x - self.dw0x
                    y = y - self.dw0y
                    for i in range(len(self.quyulist)):
                        if i % 2 == 0:
                            self.quyulist[i] = self.quyulistbak[i] + x
                        else:
                            # print(self.quyulist[i])
                            self.quyulist[i] = self.quyulistbak[i] + y
                    if len(self.quyulist1) > 4:
                        for i in range(len(self.quyulist1)):
                            if i % 2 == 0:
                                self.quyulist1[i] = self.quyulist1bak[i] - x
                            else:
                                self.quyulist1[i] = self.quyulist1bak[i] - y
                    # print(self.quyulist1)

                    break

    # 答案列表、考号、选项区域生成
    def quyulistsc(self):
        self.list1.clear()
        # self.list.clear()
        # 打开excel文件及工作表
        try:
            self.chengjiexcel = openpyxl.load_workbook(self.openexcelname)
            self.shyijuan = self.chengjiexcel['yijuan']

        except:
            print('读入答案时excel文件损坏！请查看文件及工作表！已读入moban.xlsx')
            self.chengjiexcel = openpyxl.load_workbook('moban.xlsx')
            self.shyijuan = self.chengjiexcel['yijuan']
        logging.debug('已打开选定excel文件和工作表。')
        # 读取excel答案到程序主界面
        for i in range(1, 22):
            exec('self.lineEdit_0{:02}.setText(self.shyijuan.cell(row=2, column=3+{}).value)'.format(i, i))
        logging.debug('已读取答案到程序主界面。')
        # 如果参数表cs存在，打开表
        # 清空原有列表
        logging.debug('准备清空考号区域')
        self.quyulist1.clear()
        logging.debug('准备清空选项区域')
        self.quyulist.clear()
        logging.debug('已清空旧的考号区域列表')
        if 'cs' in self.chengjiexcel.sheetnames:
            self.shcs = self.chengjiexcel['cs']
            # 读取excel内试卷参数到变量
            self.tishu = int(self.shcs.cell(row=1, column=2).value)
            self.zongfen = int(self.shcs.cell(row=2, column=2).value)
            self.meitifenshu = int(self.shcs.cell(row=3, column=2).value)
            self.bufendefen = int(self.shcs.cell(row=4, column=2).value)
            self.yuzhi = self.shcs.cell(row=5, column=2).value
            self.mianjibaifenbi = self.shcs.cell(row=6, column=2).value
            self.kk = (int(self.shcs.cell(row=7, column=2).value) - 1) * 16
            logging.debug('已读取参数到变量。')
            # 读取excel内参数到程序主界面
            self.lineEdit_tishu.setText(str(self.shcs.cell(row=1, column=2).value))
            self.lineEdit_zongfenshu.setText(str(self.shcs.cell(row=2, column=2).value))
            self.lineEdit_meitifenshu.setText(str(self.shcs.cell(row=3, column=2).value))
            self.lineEdit_bufendefen.setText(str(self.shcs.cell(row=4, column=2).value))
            self.lineEdit_chaxunxuesheng.setText(str(self.shcs.cell(row=7, column=2).value))
            self.lineEdit_yuzhi.setText(str(self.yuzhi))
            logging.debug('已读取参数到程序主界面。')
            # 默认试卷拼接开关值为0，不拼接试卷
            self.pinjiekaiguan = 0
            # 如果excel有拼接参数，拼接开关值为1，拼接试卷
            if self.shcs.cell(row=2, column=3).value != None:
                self.pinjiekaiguan = 1
                # 读取试卷拼接参数
                self.pj1x = int(self.shcs.cell(row=2, column=3).value)
                try:
                    self.pj1y = int(self.shcs.cell(row=2, column=4).value)
                    self.pj2x = int(self.shcs.cell(row=2, column=5).value)
                    self.pj2y = int(self.shcs.cell(row=2, column=6).value)
                    self.pj3x = int(self.shcs.cell(row=2, column=7).value)
                    self.pj3y = int(self.shcs.cell(row=2, column=8).value)
                except:
                    print('拼接坐标有误，重新设置！')
                    pass
            logging.debug('已读取拼接参数。')
            # 如果存在自动定位参数，读取自动定位参数
            if self.shcs.cell(row=9, column=3).value != None:
                self.dingweikaiguan = True
                self.dw1x = int(self.shcs.cell(row=9, column=2).value)
                self.dw1y = int(self.shcs.cell(row=9, column=3).value)
                self.dw2x = int(self.shcs.cell(row=10, column=2).value)
                self.dw2y = int(self.shcs.cell(row=10, column=3).value)
                self.dw0x = int(self.shcs.cell(row=11, column=2).value)
                self.dw0y = int(self.shcs.cell(row=11, column=3).value)
                logging.debug('已读取定位参数。')
                logging.debug(f'定位开关为{self.dingweikaiguan}。')
                logging.debug(f'定位参数{self.dw1x},{self.dw1y}。')
            else:
                self.dingweikaiguan = False
            # 如果考号区域有参数，生成考号识别区域self.quyulist1
            if self.shcs.cell(row=12, column=3).value != None:
                # 读取考号长度，考号数位个数
                self.khgs = int(self.shcs.cell(row=8, column=2).value)
                # 读取考号前两个点坐标
                khx1 = int(self.shcs.cell(row=12, column=2).value)
                khy1 = int(self.shcs.cell(row=12, column=3).value)
                khx2 = int(self.shcs.cell(row=13, column=2).value)
                khy2 = int(self.shcs.cell(row=13, column=3).value)
                # 如果考号第3、4个点坐标存在，读取坐标
                logging.debug('已读取考号区域前两个设置点坐标')

                if self.shcs.cell(row=14, column=3).value != None:
                    khy3 = int(self.shcs.cell(row=14, column=3).value)
                    khx4 = int(self.shcs.cell(row=15, column=2).value)
                    # khkuandu1=khx2-khx1
                    # 计算相邻考号列宽度（列间距）
                    khkuandu2 = khx4 - khx1
                    # khgaodu1=khy2-khy1
                    # 计算相邻考号行高度（行间距）
                    khgaodu2 = khy3 - khy1

                    # 循环生成考号区域列表，识别学生考号用
                    for khi in range(self.khgs):
                        for khj in range(10):
                            self.quyulist1.append(khx1 + khkuandu2 * khi)
                            self.quyulist1.append(khy1 + khgaodu2 * khj)
                            self.quyulist1.append(khx2 + khkuandu2 * khi)
                            self.quyulist1.append(khy2 + khgaodu2 * khj)
                    logging.debug('涂黑考号区域坐标已生成')
                # 如果考号第3、4个点坐标不存在，则是条形码，只要两点坐标加入考号列表self.quyulist1
                else:
                    self.quyulist1.append(khx1)
                    self.quyulist1.append(khy1)
                    self.quyulist1.append(khx2)
                    self.quyulist1.append(khy2)
                logging.debug('二条形码考号区域坐标已生成')
            # 如果没有考号参数，清空考号列表self.quyulist1
            # else:
            #     self.quyulist1 = []
            # 自动生成选项区域self.quyulist
            if self.shcs.cell(row=16, column=3).value != None:
                # 设置选项采集点列表
                listxxdian = []
                # 读取数据加到选项采集点列表
                for i in range(self.shcs.max_row - 12):
                    if self.shcs.cell(row=16 + i, column=2).value != None:
                        listxxdian.append((int(self.shcs.cell(row=16 + i, column=2).value),
                                           int(self.shcs.cell(row=16 + i, column=3).value)))
                # 如果选项ABCD横向排列
                if abs(listxxdian[2][0] - listxxdian[0][0]) > 10:
                    # 计算涂黑方块宽度
                    xxkuandu1 = listxxdian[1][0] - listxxdian[0][0]
                    # 计算ABCD选项间距
                    xxkuandu2 = listxxdian[2][0] - listxxdian[0][0]
                    # 计算黑块高度
                    xxgaodu1 = listxxdian[1][1] - listxxdian[0][1]
                    # xxgaodu2 = listxxdian[3][1] - listxxdian[0][1]
                    # print(listxxdian)
                    # print(xxkuandu1,xxkuandu2,xxgaodu1)
                    # 删除第2、3个点坐标（A的右下角、B的左上角）
                    del listxxdian[1:3]

                    # 生成新的选项区或列表
                    for xxi in listxxdian:
                        for xxj in range(4):
                            self.quyulist.append(xxi[0] + xxkuandu2 * xxj)
                            self.quyulist.append(xxi[1])
                            self.quyulist.append((xxi[0] + xxkuandu1) + xxkuandu2 * xxj)
                            self.quyulist.append(xxi[1] + xxgaodu1)
                    # print(self.quyulist)
                # 如果选项ABCD纵向排列
                else:
                    # 计算涂黑方块宽度
                    xxkuandu1 = listxxdian[1][0] - listxxdian[0][0]
                    # 计算ABCD选项间距
                    xxkuandu2 = listxxdian[2][1] - listxxdian[0][1]
                    # 计算黑块高度
                    xxgaodu1 = listxxdian[1][1] - listxxdian[0][1]
                    # xxgaodu2 = listxxdian[3][1] - listxxdian[0][1]
                    # print(listxxdian)
                    # print(xxkuandu1,xxkuandu2,xxgaodu1)
                    # 删除第2、3个点坐标（A的右下角、B的左上角）
                    del listxxdian[1:3]
                    # # 清除旧的选项区域列表
                    # self.quyulist.clear()
                    # 生成新的选项区或列表
                    for xxi in listxxdian:
                        for xxj in range(4):
                            self.quyulist.append(xxi[0])
                            self.quyulist.append(xxi[1] + xxkuandu2 * xxj)
                            self.quyulist.append((xxi[0] + xxkuandu1))
                            self.quyulist.append(xxi[1] + xxgaodu1 + xxkuandu2 * xxj)
                    # print(self.quyulist)
                    logging.debug('选择题区域坐标已生成')
        if self.dingweikaiguan == True:
            self.quyulistbak = self.quyulist.copy()
            self.quyulist1bak = self.quyulist1.copy()
            logging.debug('区域已拷贝')
        if 'gz' in self.chengjiexcel.sheetnames:
            self.shgz = self.chengjiexcel['gz']
            if self.shgz.cell(row=1, column=1).value != None:
                logging.debug('已打开gz表。')
                for i in range(self.shgz.max_column):  # 读取GZ人集合
                    self.list1.add(self.shgz.cell(row=1, column=i + 1).value)
                    # self.list11.append(self.shgz.cell(row=1, column=i + 1).value)
                # print(self.list1)
        # 读取条形码最大值最小值
        txmmin = int(self.shyijuan.cell(row=3, column=1).value)
        txmmax = int(self.shyijuan.cell(row=self.shyijuan.max_row, column=1).value)
        self.zimu = [txmmin, txmmax]
        print(self.zimu)
        # 关闭excel文件
        self.chengjiexcel.close()
        # 清除旧答案列表self.answer
        self.answer.clear()
        # 新的答案加入答案列表self.answer
        for i in range(1, 22):
            exec('self.answer.append(self.lineEdit_0{:02}.text())'.format(i))
        logging.debug('答案列表self.answer已生成')
        # 改变答案文本框样式为灰色，设置只读状态
        for i in range(1, 22):
            exec('self.lineEdit_0{:02}.setStyleSheet("background-color:rgb(230, 230, 230)")'.format(i))
            exec('self.lineEdit_0{:02}.setReadOnly(True)'.format(i))
        self.lineEdit_tishu.setStyleSheet("background-color:rgb(230, 230, 230)")
        self.lineEdit_tishu.setReadOnly(True)
        self.lineEdit_zongfenshu.setStyleSheet("background-color:rgb(230, 230, 230)")
        self.lineEdit_zongfenshu.setReadOnly(True)
        self.lineEdit_meitifenshu.setStyleSheet("background-color:rgb(230, 230, 230)")
        self.lineEdit_meitifenshu.setReadOnly(True)
        self.lineEdit_bufendefen.setStyleSheet("background-color:rgb(230, 230, 230)")
        self.lineEdit_bufendefen.setReadOnly(True)
        logging.debug('答案文本框已只读')
        # 切片题目数量
        self.answer = self.answer[: self.tishu]
        self.chengjiexcel.close()

    # 遍历文件夹中每个文件名形成列表
    def bianlipic(self):
        self.filelist.clear()
        for fname in [x for x in (Path.cwd() / 'pic').iterdir()]:
            self.filelist.append(fname)
        if len(self.filelist) != 0:
            print('试卷列表读取成功！')
            return self.filelist
        else:
            print('pic内无试卷！')

    # 处理图为二值图
    def picchuli(self, path):
        # BGR模式读取，写入试卷，才可能用红色写入
        self.img = cv2.imread(str(path))
        # 灰度图用于处理试卷
        huidu = cv2.imread(str(path), 0)
        logging.debug("试卷读入转化为灰度正常")
        # 反向二值化
        t, erzhihua = cv2.threshold(huidu, int(self.yuzhi), 255, cv2.THRESH_BINARY_INV)
        logging.debug("试卷二值化处理正常")

        self.kaiyunsuan = erzhihua

        logging.debug(f'kg:{self.dingweikaiguan}')
        # 允许自动调是开关值为1
        if self.dingweikaiguan == True:
            logging.debug('准备自动调整区域')
            self.quyuauto()  # 自动调整区域

    # 考号识别
    def kaohaoshibie(self):
        # 识别成功的考号
        self.kauanohaoch = []
        # 识别出的考号单个字符
        self.kaohao = []
        if len(self.quyulist1) > 4:
            logging.debug('开始识9位考号')
            # logging.debug(self.quyulist1)
            kk = 0
            # 循环识别多位考号
            for ii in range(self.khgs):  # 识别9位考号
                jjlist = []
                jjbiliulist = []
                # 循环识别每位考号的10个数字
                for jj in range(0, 40, 4):
                    x1 = int(self.quyulist1[jj + kk])
                    x2 = int(self.quyulist1[jj + kk + 2])
                    y1 = int(self.quyulist1[jj + kk + 1])
                    y2 = int(self.quyulist1[jj + kk + 3])
                    hk1 = self.kaiyunsuan[y1:y2, x1:x2]
                    # 计算白色面积
                    baisemianji = cv2.countNonZero(hk1)
                    # 计算一位数字区域的总面积
                    quanbumianji = (self.quyulist1[2] - self.quyulist1[0]) * (
                            self.quyulist1[3] - self.quyulist1[1])
                    # 计算涂黑区域占总区域面积百分比
                    ratio = baisemianji * 100 / quanbumianji
                    # 如果大于指定比率，加入索引列表和比率列表
                    if ratio > self.mianjibaifenbi:
                        jjlist.append(jj)
                        jjbiliulist.append(jjbiliulist)
                if len(jjbiliulist) > 0:  # 如果有识别出的数字
                    # 返回10位数字最大占比的索引值
                    jj = jjlist[jjbiliulist.index(max(jjbiliulist))]
                    # 数字中最在比率的为识别结果，加入self.kaohao
                    if jj == 0:
                        self.kaohao.append('0')
                    elif jj == 4:
                        self.kaohao.append('1')
                    elif jj == 8:
                        self.kaohao.append('2')
                    elif jj == 12:
                        self.kaohao.append('3')
                    elif jj == 16:
                        self.kaohao.append('4')
                    elif jj == 20:
                        self.kaohao.append('5')
                    elif jj == 24:
                        self.kaohao.append('6')
                    elif jj == 28:
                        self.kaohao.append('7')
                    elif jj == 32:
                        self.kaohao.append('8')
                    elif jj == 36:
                        self.kaohao.append('9')
                    kk = kk + 40
                else:  # 如果没有识别出的数字
                    print('该生没涂考号！')
                    break

            # 列表变为字符串
            self.kauanohaoch = ''.join(self.kaohao)
            print('识别考号：' + str(self.kauanohaoch))
        elif len(self.quyulist1) == 4:
            logging.debug('开始识别条形码')
            # logging.debug(self.quyulist1)
            # 从条形码区域列表中提取条形码位置坐标。
            min_x = min(int(self.quyulist1[0]), int(self.quyulist1[2]))
            min_y = min(int(self.quyulist1[1]), int(self.quyulist1[3]))
            width = abs(int(self.quyulist1[0]) - int(self.quyulist1[2]))
            height = abs(int(self.quyulist1[1]) - int(self.quyulist1[3]))
            # 设置条形码所在区域

            cut_img = self.img[min_y:min_y + height, min_x:min_x + width]
            lslist88 = pyzbar.decode(cut_img)
            if len(lslist88) == 1:
                # 识别考号
                self.kauanohaoch = lslist88[0].data.decode("utf-8")

                print('识别考号：' + str(self.kauanohaoch))
            else:
                self.tiaoxingmacw = self.tiaoxingmacw + str(self.filelist[self.filenumjsq])[-6:]
                print('没有找到条形码！请检查目录文件夹或条形码区域是否正确！')
        else:
            print('没有考号！')

    # 选择题选项识别
    def shitipanbie(self):
        self.jianquekg = True
        # 字典键名
        self.key = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21]
        # 学生选项字典（判卷中使用）
        self.charuxuan = dict([(k, []) for k in self.key])
        # 学生选项字典（判卷后使用）
        self.gengxinruxuan = dict([(k, []) for k in self.key])
        # 学生选项列表
        self.xuan = []
        # kk为选区起点，10为从11题开始读卡。0*16为从1题读卡
        kk = 0
        # 识别选择题 ii为题号索引
        for ii in range(self.tishu):
            logging.debug('开始识别第{}题'.format(ii + 1))
            for jj in range(0, 16, 4):
                # 取出每个选项区域坐标
                x1 = int(self.quyulist[jj + kk])
                x2 = int(self.quyulist[jj + kk + 2])
                y1 = int(self.quyulist[jj + kk + 1])
                y2 = int(self.quyulist[jj + kk + 3])
                logging.debug('开始识别{},{},{},{}项'.format(x1, y1, x2, y2))
                hk = self.kaiyunsuan[y1:y2, x1:x2]
                # 计算涂黑面积
                baisemianji = cv2.countNonZero(hk)

                # 计算选项总的面积
                quanbumianji = (self.quyulist[2] - self.quyulist[0]) * (self.quyulist[3] - self.quyulist[1])
                logging.debug('面积各为：{}{}'.format(baisemianji, quanbumianji))
                # 计算涂黑面积占选项区域的总面积的百分比
                ratio = baisemianji * 100 / quanbumianji
                logging.debug('比率为：{}'.format(ratio))
                logging.debug('jj：{}'.format(jj))
                # 如果占比超过设置，加入选项字典self.charuxuan以题号为key的选项列表
                if ratio > self.mianjibaifenbi:
                    if jj == 0:
                        logging.debug(f'第 {ii + 1} 题选：A {ratio}')
                        self.charuxuan[ii + 1].append('A')
                    elif jj == 4:
                        self.charuxuan[ii + 1].append('B')
                        logging.debug(f'第 {ii + 1} 题选：B{ratio} ')
                    elif jj == 8:
                        self.charuxuan[ii + 1].append('C')
                        logging.debug(f'第 {ii + 1} 题选：C {ratio}')
                    elif jj == 12:
                        self.charuxuan[ii + 1].append('D')
                        logging.debug(f'第 {ii + 1} 题选：D{ratio} ')
            if (len(self.answer[ii]) == 1) and (len(self.charuxuan[ii + 1]) > 1):
                logging.debug('进入单选误判处理环节')
                self.duoxuanset.add((str(self.filelist[self.filenumjsq]))[-6:])
                print(f'检查第{ii + 1}题区域是否对齐，单选题测出多个选项。')
                self.duoxuanset.add((str(self.filelist[self.filenumjsq]))[-6:])
                self.jianquekg = False
                if self.lineEdit_chaxunxuesheng.text() == '3':
                    self.jianquekg = True

            if len(self.charuxuan[ii + 1]) == 4:
                self.duoxuanset.add((str(self.filelist[self.filenumjsq]))[-6:])
                self.jianquekg = False
                if self.lineEdit_chaxunxuesheng.text() == '3':
                    self.jianquekg = True
            # 字典self.charuxuan的选项列表转化为字符串加入字典self.gengxinruxuan题号对应值的列表中
            self.gengxinruxuan[ii + 1].append(''.join(self.charuxuan[ii + 1]))
            kk = kk + 16
        # 字典变成二维列表
        self.xuan = list(self.gengxinruxuan.values())
        # 二维列表变成一维
        self.xuan = list(chain.from_iterable(self.xuan))
        # 切片除题目数量
        self.xuan = self.xuan[: self.tishu]
        print('正确答案是：', self.answer)
        print('学生答案是：', self.xuan)

    # 计算一份试卷答题情况并打印在试卷上，做错的圈出正确答案
    def tongjitdayin(self):
        # 空白选项的试卷
        self.kongxuanxian = set()

        # 答对题目数量
        self.right = 0
        # 半对题目数量
        self.bandui = 0
        # 空白题目数量
        self.kong = 0
        # 遍历每个正确答案和学生答案
        for i in range(len(self.answer)):
            # 取出学生答案和标准答案转化为集合
            dd = set(self.answer[i].upper())
            xx = set(self.xuan[i].upper())
            # 如果正确，正确数目加1
            if xx == dd and self.answer[i] != '':
                self.right = self.right + 1
            # 如果半对，半对数目加1且题目上打上扣除分数
            elif (xx.issubset(dd) == True) and (dd.difference(xx) != set()) and (self.answer[i] != '') and (
                    self.xuan[i] != ''):
                self.bandui = self.bandui + 1
                # 如果是竖卡（标准机读卡）分数打在题号下
                if self.img.shape[0] < self.img.shape[0]:
                    cv2.putText(self.img, '-' + str(self.meitifenshu - self.bufendefen),
                                (self.quyulist[16 * i], self.quyulist[16 * i + 1]),
                                cv2.FONT_HERSHEY_TRIPLEX, 1, (255, 0, 0), 2)
                # 如果是横卡（二卷答题卡）分数打在选项后
                else:
                    cv2.putText(self.img, '-' + str(self.meitifenshu - self.bufendefen),
                                (self.quyulist[16 * i + 14], self.quyulist[16 * i + 1 + 12]),
                                cv2.FONT_HERSHEY_TRIPLEX, 1, (0, 0, 255), 1)
            # 如果答错
            else:
                # 如果是竖卡（标准机读卡）分数打在题号下
                if self.img.shape[0] < self.img.shape[0]:
                    cv2.putText(self.img, '-' + str(self.meitifenshu),
                                (self.quyulist[16 * i], self.quyulist[16 * i + 1]),
                                cv2.FONT_HERSHEY_TRIPLEX, 1, (255, 0, 0), 2)
                # 如果是横卡（二卷答题卡）分数打在选项后
                else:
                    cv2.putText(self.img, '-' + str(self.meitifenshu),
                                (self.quyulist[16 * i + 14], self.quyulist[16 * i + 1 + 12]),
                                cv2.FONT_HERSHEY_TRIPLEX, 1, (0, 0, 255), 1)
                # 如果选项为空，self.kong+1
                if xx == '':
                    self.kong = self.kong + 1
                    # 试卷名放入字符串，用于阅卷完毕时提醒
                if '' in self.xuan:
                    self.kongxuanxian.add(self.name + str(self.filelist[self.filenumjsq]).split('\\')[-1])
        if self.xuan.count('') >= 1:
            self.jianquekg = False
            if self.lineEdit_chaxunxuesheng.text() == '3':
                self.jianquekg = True

            self.kongxuanxiansum = self.kongxuanxiansum + str(self.kongxuanxian)
        logging.debug("扣分打印正常")
        # 计算学生总分
        strfs = str(self.right * self.meitifenshu + self.bandui * self.bufendefen)
        print(strfs + '分')
        # 试卷打印上总分
        cv2.putText(self.img, '{} '.format(strfs),
                    (self.quyulist[-2], self.quyulist[-1] + 60), cv2.FONT_HERSHEY_TRIPLEX, 1, (0, 0, 255), 2)
        logging.debug("总分打印正常")
        # 红框圈住答题的选项
        if self.jianquekg == True:
            for i in range(self.tishu):
                if self.answer[i] != self.xuan[i]:
                    if 'A' in self.answer[i]:
                        x1 = int(self.quyulist[i * 16])
                        x2 = int(self.quyulist[i * 16 + 2])
                        y1 = int(self.quyulist[i * 16 + 1])
                        y2 = int(self.quyulist[i * 16 + 3])
                        if self.img.shape[0] < self.img.shape[0]:
                            cv2.rectangle(self.img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                        else:
                            cv2.rectangle(self.img, (x1, y1), (x2, y2), (0, 0, 255), 2)
                    if 'B' in self.answer[i]:
                        x1 = int(self.quyulist[i * 16 + 4])
                        x2 = int(self.quyulist[i * 16 + 6])
                        y1 = int(self.quyulist[i * 16 + 5])
                        y2 = int(self.quyulist[i * 16 + 7])
                        if self.img.shape[0] < self.img.shape[0]:
                            cv2.rectangle(self.img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                        else:
                            cv2.rectangle(self.img, (x1, y1), (x2, y2), (0, 0, 255), 2)
                    if 'C' in self.answer[i]:
                        x1 = int(self.quyulist[i * 16 + 8])
                        x2 = int(self.quyulist[i * 16 + 10])
                        y1 = int(self.quyulist[i * 16 + 9])
                        y2 = int(self.quyulist[i * 16 + 11])
                        if self.img.shape[0] < self.img.shape[0]:
                            cv2.rectangle(self.img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                        else:
                            cv2.rectangle(self.img, (x1, y1), (x2, y2), (0, 0, 255), 2)
                    if 'D' in self.answer[i]:
                        x1 = int(self.quyulist[i * 16 + 12])
                        x2 = int(self.quyulist[i * 16 + 14])
                        y1 = int(self.quyulist[i * 16 + 13])
                        y2 = int(self.quyulist[i * 16 + 15])
                        if self.img.shape[0] < self.img.shape[0]:
                            cv2.rectangle(self.img, (x1, y1), (x2, y2), (0, 255, 0), 2)
                        else:
                            cv2.rectangle(self.img, (x1, y1), (x2, y2), (0, 0, 255), 2)
            cv2.imwrite(str(self.filelist[self.filenumjsq]), self.img)

    # 成绩写入excel
    def xiechengji(self):
        # 1打开excel文件及表yijuan
        try:
            self.chengjiexcel = openpyxl.load_workbook(self.openexcelname)  # 打开excel
        except:
            print('excel文件损坏！')
        logging.debug("已打开文件{}".format(self.openexcelname))
        self.shyijuan = self.chengjiexcel['yijuan']
        logging.debug("已打开文件yijuan表")
        # 2如果有考号，进行有考号写成绩模式
        if len(self.quyulist1) > 0:
            logging.debug("进入有考号处理模式")
            # （1）遍历excel考号与试卷识别考号对比，如果相等则进入写成绩模式
            for i in range(3, self.shyijuan.max_row + 1):
                self.xuehao = self.shyijuan.cell(row=i, column=1).value
                if (str(self.xuehao)) == str(self.kauanohaoch):
                    # 4选择题识别
                    self.shitipanbie()
                    # 读取学生姓名
                    self.name = self.shyijuan.cell(row=i, column=2).value
                    if self.jianquekg == True:
                        # 成绩计算，错题打印标记
                        self.tongjitdayin()

                        # 读取excel选择题成绩
                        fenshu = self.shyijuan.cell(row=i, column=3).value
                        # 计算所有试卷总分
                        self.zongfenbj = self.zongfenbj + self.right * self.meitifenshu + self.bandui * self.bufendefen

                        # 学生选项写入excel
                        for xx2 in range(len(self.answer)):
                            self.shyijuan.cell(row=i, column=4 + xx2, value=self.xuan[xx2])

                        # 写入I卷总分
                        self.shyijuan.cell(row=i,
                                           column=3).value = self.right * self.meitifenshu + self.bandui * self.bufendefen
                        # 批改成功的试卷路径存入excel
                        self.shyijuan.cell(row=i, column=len(self.answer) + 4).value = str(
                            self.filelist[self.filenumjsq]).replace('pic', 'bak')
                        # 以下四行把成功识别的试卷放入bak下
                        if self.jianquekg == True:
                            # 有效试卷读数加1
                            self.youxiaonum = self.youxiaonum + 1
                            p = Path.cwd()
                            p2 = Path(str(self.filelist[self.filenumjsq]))
                            target = p.joinpath('bak').joinpath(p2.name)
                            if target.exists():
                                target.unlink()
                            p2.rename(target)
                        print('姓名：', self.name, '分数：',
                              self.right * self.meitifenshu + self.bandui * self.bufendefen)

            try:
                self.chengjiexcel.save(self.openexcelname)
            except:
                print('保存出错，先关闭excel再重试！')
            self.chengjiexcel.close()
        # 没有考号时，按照试卷顺序写入成绩
        else:
            logging.debug("进入无考号处理模式")
            # 4选择题识别
            self.shitipanbie()
            # 成绩计算，错题打印标记
            self.tongjitdayin()
            # 文件读数器+3好与入的行数i,一份试卷只写一行，所以一次只循环一行
            for i in [self.filenumjsq + 3, self.filenumjsq + 4]:
                logging.debug(self.filenumjsq + 3)
                # 读取I卷总分
                fenshu = self.shyijuan.cell(row=i, column=3).value
                # 计算所有试卷总分
                self.zongfenbj = self.zongfenbj + self.right * self.meitifenshu + self.bandui * self.bufendefen
                # 本行控制：允许空白I卷分时写入、更新成绩时必须为同一张试卷且分数变大才能写入
                if fenshu == None or (
                        int(fenshu) < int(self.right * self.meitifenshu + self.bandui * self.bufendefen) and (
                        self.shyijuan.cell(row=i, column=len(self.answer) + 4).value.split('\\')[-1] ==
                        str(self.filelist[self.filenumjsq]).split('\\')[-1])):

                    logging.debug('进入写成绩模式')
                    # 如果原来有成绩更新了，提示已更新
                    if fenshu != None:
                        # 获取试卷文件名
                        yuanwenjuan = self.shyijuan.cell(row=i, column=len(self.answer) + 4).value.split('\\')[-1]
                        print(yuanwenjuan, fenshu, '分，更新为', str(self.filelist[self.filenumjsq]).split('\\')[-1],
                              self.right * self.meitifenshu + self.bandui * self.bufendefen, '分!!!!!!')
                    logging.debug('已写入成绩')
                    # 写入学生每题选项
                    for xx2 in range(len(self.answer)):
                        self.shyijuan.cell(row=i, column=4 + xx2, value=self.xuan[xx2])
                    logging.debug('已写入选项')

                    # 写入I卷总分
                    self.shyijuan.cell(row=i,
                                       column=3).value = self.right * self.meitifenshu + self.bandui * self.bufendefen
                    logging.debug('写入总分')
                    # 批改成功的试卷修改路径存入excel
                    self.shyijuan.cell(row=i, column=len(self.answer) + 4).value = str(
                        self.filelist[self.filenumjsq]).replace('pic', 'bak')
                    # 以下四行把成功识别的试卷放入bak下
                    if self.jianquekg == True:
                        # 批改成功有效试卷读数器加1
                        self.youxiaonum = self.youxiaonum + 1
                        logging.debug('有效试卷份数加1')
                        p = Path.cwd()
                        p2 = Path(str(self.filelist[self.filenumjsq]))
                        target = p.joinpath('bak').joinpath(p2.name)
                        if target.exists():
                            target.unlink()
                        p2.rename(target)
                        logging.debug('已移动试卷到bak下')
                    print('分数：', self.right * self.meitifenshu + self.bandui * self.bufendefen)
                else:
                    print('没有写入成绩，分数已存在！')

            try:
                self.chengjiexcel.save(self.openexcelname)
            except:
                print('保存出错，先关闭excel再重试！')
            self.chengjiexcel.close()

    # 开始阅卷按钮
    @pyqtSlot()
    def on_pushButton_kaishi_clicked(self):
        # 试卷文件列表指针
        self.filenumjsq = 0
        # 批阅完成试卷计数
        self.youxiaonum = 0
        logging.debug('self.youxiao正确')
        # 选项中有空白试卷文件名
        self.kongxuanxiansum = ''
        # 有效试卷总分
        self.zongfenbj = 0
        # 试卷列表清空
        self.filelist = []
        self.bianlipic()
        # 打印常规信息，提供出错诊断
        print('试卷份数：', len(self.filelist))
        self.duoxuanset.clear()
        self.tiaoxingmacw = ''
        if len(self.filelist) > 0:
            print('试卷格式为：' + str(self.filelist[self.filenumjsq])[-3:])
            if str(self.filelist[self.filenumjsq])[-3:] != 'jpg':
                print('图片格式不是jpg!')
                # sys.exit(0)
        print('excel文件是：', self.openexcelname)
        print('试题个数与答案个数为：', self.tishu, len(self.answer))
        print('试卷文件夹为：', self.wenjianjia)
        for img in glob.glob("ls/*.*"):
            os.remove(img)
        logging.debug("清理ls目录成功")
        # 1遍历试卷文件名，形成列表

        for i in range(len(self.filelist)):
            # 2试卷二值化处理
            self.picchuli(self.filelist[self.filenumjsq])
            print('*' * 60)
            print('正在识别：', self.filelist[self.filenumjsq])
            # 3.考号识别
            try:
                self.kaohaoshibie()
                # 5成绩存入excel
                if self.tiaoxuanshijuankaiguan == False:
                    self.xiechengji()
                    self.filenumjsq = self.filenumjsq + 1
                    if self.lineEdit_chaxunxuesheng.text() == '2':
                        self.jianquekg = True
                        # 本行用于查看选区是否正确
                        self.quqyu_huatu()
                        # 调试时打开以下三行，可以看出二值化开运算结果问题
                        logging.debug('区域画图没问题')
                        # 0可以改变窗口大小位置，缺省则只能移动，不能改大小
                        cv2.namedWindow("22", 0)
                        cv2.resizeWindow("22", 1000, 1000)
                        cv2.imshow('22', self.kaiyunsuan)
                        cv2.waitKey()
                        cv2.destroyAllWindows()

                else:
                    if self.zimu[1] >= int(self.kauanohaoch) >= self.zimu[0]:
                        print('是需要的试卷')
                        # movetols
                        os.rename(self.filelist[self.filenumjsq],
                                  str(self.filelist[self.filenumjsq]).replace('pic', 'ls'))
                        os.rename(self.filelist[self.filenumjsq + 1],
                                  str(self.filelist[self.filenumjsq + 1]).replace('pic', 'ls'))
                        self.filenumjsq = self.filenumjsq + 2
                    else:
                        print('不是需要的试卷')
                        os.remove(self.filelist[self.filenumjsq])
                        os.remove(self.filelist[self.filenumjsq + 1])
                        self.filenumjsq = self.filenumjsq + 2

                # 6总体阅卷情况
                if self.filenumjsq >= len(self.filelist):
                    print('*****已阅' + str(self.filenumjsq) + '份,录入' + str(self.youxiaonum) + '份')
                    break
            except:
                self.tiaoxingmacw = self.tiaoxingmacw + str(
                    '条形码错误！请处理完{}再运行程序。'.format(self.filelist[self.filenumjsq]))
                self.filenumjsq = self.filenumjsq + 1
                if self.filenumjsq >= len(self.filelist):
                    # if self.youxiaonum > 0:
                    print('*****已阅' + str(self.filenumjsq) + '份,录入' + str(self.youxiaonum) + '份')

                    # else:
                    #     print('*****已阅' + str(self.filenumjsq) + '份,录入' + str(self.youxiaonum) + '份')
                    break
                continue

        self.tiaoxuanshijuankaiguan = False
        self.lineEdit_yuzhi.setText(str(self.yuzhi))
        if len(self.kongxuanxiansum):
            print(self.kongxuanxiansum, '选项有空白，请检查试卷。')
        if len(self.duoxuanset) > 0:
            print(self.duoxuanset, '单选题出现多选，请提高比率。')
        if len(self.tiaoxingmacw) > 0:
            print(self.tiaoxingmacw, '这些试卷条形码错误。')

    # 考号选项区域打印在试卷上，加载到标签中
    def quqyu_huatu(self):
        # 清空文件列表
        self.filelist.clear()
        # 获取pic路径对象
        imagePath = Path(sys.argv[0]).parent
        imagePath2 = imagePath.joinpath('pic')
        # 遍历文件夹中每个文件名形成列表
        for fname in [x for x in imagePath2.iterdir()]:
            self.filelist.append(fname)
        if len(self.filelist) != 0:
            print('试卷列表读取成功！')
        else:
            print('pic内无试卷！')
        # 试卷画上考号和选项区域
        try:
            # 读取pic内试卷为彩色模式
            img = cv2.imread(str(self.filelist[self.filenumjsq - 1]))
            logging.debug('试卷{}已打开'.format(str(self.filelist[self.filenumjsq - 1])))
            logging.debug('列表{}'.format(self.filelist))
            logging.debug('计数器{}'.format(self.filenumjsq))
            # 画出选项区域
            for i in range(0, len(self.quyulist), 4):  # 选项
                img = cv2.rectangle(img, (self.quyulist[i], self.quyulist[i + 1]),
                                    (self.quyulist[i + 2], self.quyulist[i + 3]), (0, 0, 255), 2)
            logging.debug('选项区域画图成功')
            # 如果有考号区域，画出考号区域
            if len(self.quyulist1) > 0:
                for i in range(0, len(self.quyulist1), 4):
                    img = cv2.rectangle(img, (self.quyulist1[i], self.quyulist1[i + 1]),
                                        (self.quyulist1[i + 2], self.quyulist1[i + 3]), (0, 0, 255), 2)
            logging.debug('考号区域画图成功')
            # 标签加载图片
            shrink = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
            QtImg = QImage(shrink.data,
                           shrink.shape[1],
                           shrink.shape[0],
                           shrink.shape[1] * 3,
                           QImage.Format_RGB888)
            self.label_shijuan.setPixmap(
                QPixmap.fromImage(QtImg).scaled(self.label_shijuan.width(), self.label_shijuan.height()))
            logging.debug('加载图像成功')

        except:
            print('pic下没有试卷！')

    # 调整区域
    @pyqtSlot()
    def on_pushButton_tiaozhengquyu_clicked(self):
        # print(self.quyulist)
        self.dingweikaiguan = False  # 手动调整区域，则关闭自动调整
        listinpu = (self.lineEdit_chaxunxuesheng.text()).split()
        self.filenumjsq = 1
        if len(listinpu) == 3:  # 所有区域都调整
            x, y, z = listinpu
            x = int(x)
            y = int(y)
            for i in range(len(self.quyulist)):
                if i % 2 == 0:
                    self.quyulist[i] = self.quyulist[i] + x
                else:
                    self.quyulist[i] = self.quyulist[i] + y
            for i in range(len(self.quyulist1)):
                if i % 2 == 0:
                    self.quyulist1[i] = self.quyulist1[i] + x
                else:
                    self.quyulist1[i] = self.quyulist1[i] + y
        elif len(listinpu) == 2:  # 只调选项区域
            logging.debug('只调选区域')
            x, y = listinpu
            x = int(x)
            y = int(y)
            for i in range(len(self.quyulist)):
                if i % 2 == 0:
                    self.quyulist[i] = self.quyulist[i] + x
                else:
                    self.quyulist[i] = self.quyulist[i] + y
        else:
            pass
        # print(self.quyulist)
        self.quqyu_huatu()

    # 导出excel成绩
    @pyqtSlot()
    def on_pushButton_daochuexcel_clicked(self):
        # 打开excel
        try:
            self.chengjiexcel = openpyxl.load_workbook(self.openexcelname, data_only=True)
            logging.debug("已打开文件{}".format(self.openexcelname))
        except:
            print('excel文件损坏！')
        # 读取第一个文本框内容设为导出文件名
        shijuanlaiyuan = self.lineEdit_chaxunxuesheng.text()
        fontred1 = Font(u'等线', size=11, bold=False, italic=False, strike=False, color='FF0000')
        print(shijuanlaiyuan)
        logging.debug('已获取试卷来源{}'.format(shijuanlaiyuan))
        # 打开表单
        self.shyijuan = self.chengjiexcel['yijuan']
        logging.debug('已打开工作表yijuan')
        self.shtongjish = self.chengjiexcel['tongjish']
        logging.debug('已打开工作表tongjis')
        # 如果一卷前2列数据不为空，统计一卷成绩。
        if (self.shyijuan.cell(row=3, column=3).value is not None) or (
                self.shyijuan.cell(row=4, column=3).value is not None):
            # 对每一列即每一题进行统计
            for kk in range(self.tishu):
                logging.debug('开始统计第{}题'.format(kk + 1))
                listright = set()
                strall = ''
                listall = set()
                daan = self.shyijuan.cell(row=2, column=kk + 4).value
                # print(self.tishu)
                for i in range(3, self.shyijuan.max_row + 1):  # 从第2行开始，到第3行结束

                    if self.shyijuan.cell(row=i, column=kk + 4).value is not None:
                        # 把选项不为空的人名加入总人数集合listall
                        listall.add(self.shyijuan.cell(row=i, column=2).value)
                        if self.shyijuan.cell(row=i, column=kk + 4).value == daan:
                            # 该题答对人名加入集合listright
                            listright.add(self.shyijuan.cell(row=i, column=2).value)
                        # print(listright)
                        # 把选项加入选项字符串
                        strall = strall + self.shyijuan.cell(row=i, column=kk + 4).value
                logging.debug('已统计总人数：{},答对人数：{}'.format(len(listall), len(listright)))
                self.shtongjish.cell(row=4, column=kk + 1).value = '正答率' + str(
                    int((len(listright) / len(listall)) * 100)) + '%'
                if int((len(listright) / len(listall)) * 100) < 50:
                    self.shtongjish.cell(row=4, column=kk + 1).font = fontred1
                self.shtongjish.cell(row=5, column=kk + 1).value = '选A' + str(
                    int((strall.count('A')) / len(listall) * 100)) + '%'
                self.shtongjish.cell(row=6, column=kk + 1).value = '选B' + str(
                    int((strall.count('B')) / len(listall) * 100)) + '%'
                self.shtongjish.cell(row=7, column=kk + 1).value = '选C' + str(
                    int((strall.count('C')) / len(listall) * 100)) + '%'
                self.shtongjish.cell(row=8, column=kk + 1).value = '选D' + str(
                    int((strall.count('D')) / len(listall) * 100)) + '%'

                if len(self.list1) > 0:
                    font3 = Font(u'微软雅黑', size=8, bold=False, italic=False, strike=False, color='ffffff')  # 设置字体样式
                    self.shtongjish.cell(row=9, column=kk + 1,
                                         value=''.join(list((listall - listright) & self.list1))).font = font3
                logging.debug('已统计正答率')

            logging.debug('所有题正答案率都统计完毕')
            manfenren = ''
            bufenfenren2 = ''
            bufenfenren = ''
            renfenzhidian = dict()
            renfenzhidian2 = dict()
            qianshimingdan = ''
            # 对总分进行统计
            for kk in range(self.shyijuan.max_row):
                if self.shyijuan.cell(row=kk + 3, column=3).value != None:
                    # print(self.shyijuan.cell(row=kk + 3, column=2).value,int(self.shyijuan.cell(row=kk + 3, column=3).value))
                    renfenzhidian[self.shyijuan.cell(row=kk + 3, column=2).value] = int(
                        self.shyijuan.cell(row=kk + 3, column=3).value)
                    # print(renfenzhidian)
                    if int(self.shyijuan.cell(row=kk + 3, column=3).value) == int(self.meitifenshu * self.tishu):
                        manfenren = manfenren + self.shyijuan.cell(row=kk + 3, column=2).value + ' '
                        # print('manfen',self.shyijuan.cell(row=kk+3, column=2).value)
                    if int(self.shyijuan.cell(row=kk + 3, column=3).value) == int(
                            self.meitifenshu * self.tishu - self.bufendefen):
                        bufenfenren = bufenfenren + self.shyijuan.cell(row=kk + 3, column=2).value + ' '
                        # print('bufenfen',self.shyijuan.cell(row=kk+3, column=2).value)
                    if int(self.shyijuan.cell(row=kk + 3, column=3).value) == int(
                            self.meitifenshu * self.tishu - self.bufendefen * 2):
                        bufenfenren2 = bufenfenren2 + self.shyijuan.cell(row=kk + 3, column=2).value + ' '
                        # print('bufen2',self.shyijuan.cell(row=kk+3, column=2).value)
            logging.debug('总分统计完毕')
            ls = list(renfenzhidian.items())
            # print(ls[0])
            # 对总分进行排序
            ls.sort(key=lambda x: x[1], reverse=True)
            logging.debug('总分已排序，最高分是：'.format(ls[0]))
            logging.debug("名单为{}".format(ls[0]))
            # 输出前9名
            for i in range(9):
                if len(ls) >= 9:
                    qianshimingdan = qianshimingdan + '{}{} '.format(ls[i][0], ls[i][1])
                    self.shtongjish.cell(row=15, column=2).value = qianshimingdan
            logging.debug("前9名已输出")
            qianshimingdan = ''
            # 输出9-18名
            for i in range(9, 18):
                if len(ls) >= 18:
                    qianshimingdan = qianshimingdan + '{}{} '.format(ls[i][0], ls[i][1])
                    self.shtongjish.cell(row=16, column=2).value = qianshimingdan
            logging.debug('已输出名次')
            fontred = Font(u'微软雅黑', size=9, bold=True, italic=False, strike=False, color='FF0000')
            self.shtongjish.cell(row=10, column=1, value='I卷').font = fontred
            self.shtongjish.cell(row=16, column=1).value = '10-18名：'
            self.shtongjish.cell(row=15, column=1).value = ' 1- 9名：'
            self.shtongjish.cell(row=11, column=1).value = '满分' + str(self.meitifenshu * self.tishu) + '名单：'
            self.shtongjish.cell(row=13, column=1).value = str(
                self.meitifenshu * self.tishu - self.bufendefen * 2) + '分名单：'
            self.shtongjish.cell(row=12, column=1).value = str(self.meitifenshu * self.tishu - self.bufendefen) + '分名单：'
            self.shtongjish.cell(row=12, column=2).value = bufenfenren
            self.shtongjish.cell(row=13, column=2).value = bufenfenren2
            self.shtongjish.cell(row=11, column=2).value = manfenren
        # 输出标题
        self.shijianstr1 = '{}年{}月{}日'.format(time.localtime().tm_year,
                                              time.localtime().tm_mon,
                                              time.localtime().tm_mday)
        self.shijianstr = '{}{}{}'.format(time.localtime().tm_year,
                                          time.localtime().tm_mon,
                                          time.localtime().tm_mday)
        self.shtongjish.cell(row=1, column=1).value = self.shijianstr1
        logging.debug("标题已输出")
        # 如果有二卷成绩，分析二卷
        if 'erjuan' in self.chengjiexcel.sheetnames:
            erjuan = self.chengjiexcel['erjuan']
            # 如果二卷这两个数据不为空，分析二卷
            if (erjuan.cell(row=3, column=3).value is not None) and (
                    erjuan.cell(row=4, column=3).value is not None) and (
                    erjuan.cell(row=5, column=3).value is not None):
                logging.debug("二卷已打开")
                rows = erjuan.max_row
                cols = erjuan.max_column
                logging.debug('行数{}列数{}'.format(rows, cols))
                # 计算试卷总分写入excel
                for i in range(3, rows + 1):
                    ejzf = 0
                    for j in range(2, cols):
                        # logging.debug('1卷')
                        if erjuan.cell(row=i, column=j).value != None:
                            ejzf = int(erjuan.cell(row=i, column=j).value) + ejzf
                    logging.debug(ejzf)
                    if self.shyijuan.cell(row=i, column=3).value != None:
                        erjuan.cell(row=i, column=cols).value = int(self.shyijuan.cell(row=i, column=3).value) + ejzf
                    else:
                        erjuan.cell(row=i, column=cols).value = ejzf
                    logging.debug(i)
                if cols > 1:
                    logging.debug('00')
                    self.shtongjish.cell(row=18, column=1, value='II卷').font = fontred
                    # self.shtongjish.cell(row=18, column=1, value='II卷')
                    logging.debug('11')
                    self.shtongjish.cell(row=19, column=1).value = '题号：'

                    self.shtongjish.cell(row=20, column=1).value = '均分：'
                    self.shtongjish.cell(row=21, column=1).value = '满分：'
                    logging.debug('12')
                    self.shtongjish.cell(row=23, column=1, value='总分').font = fontred

                    # self.shtongjish.cell(row=23, column=1, value='总分')
                    self.shtongjish.cell(row=24, column=1).value = '1-8名：'
                    self.shtongjish.cell(row=25, column=1).value = '9-19名：'
                    logging.debug('标题完毕')
                    # erjuan题号到tongji表erjuan题号
                    for i in range(2, cols):
                        self.shtongjish.cell(row=19, column=i).value = erjuan.cell(row=1, column=i).value
                    logging.debug('题号完毕')
                    # erjuan每题均分到tongji表
                    for i in range(2, cols):
                        jsq = 0
                        shitizf = 0
                        manfenmd = ''
                        mfjsq = 0
                        for j in range(3, rows + 1):
                            logging.debug(erjuan.cell(row=j, column=i).value)
                            if erjuan.cell(row=j, column=i).value != None:
                                logging.debug(jsq)
                                logging.debug(shitizf)
                                jsq = jsq + 1
                                shitizf = shitizf + int(erjuan.cell(row=j, column=i).value)
                                if erjuan.cell(row=j, column=i).value == erjuan.cell(row=2, column=i).value:
                                    manfenmd = manfenmd + self.shyijuan.cell(row=j, column=2).value
                                    mfjsq = mfjsq + 1
                        self.shtongjish.cell(row=20, column=i).value = int(shitizf / jsq)
                        logging.debug('均分统计完毕')
                        # 统计满分人数和人名
                        if mfjsq > 0:
                            self.shtongjish.cell(row=21, column=i).value = str(mfjsq) + '人:' + manfenmd
                        # 统计二卷满分人数
                        if jsq < rows - 2:
                            self.shtongjish.cell(row=22, column=i).value = '{}人'.format(jsq)
                    logging.debug('满分统计完毕')
                    # erjuan成绩临时搬到yijuan表内
                    for i in range(2, cols + 1):
                        for j in range(1, rows + 1):
                            # print(j,i)
                            self.shyijuan.cell(row=j, column=self.tishu + i + 2).value = erjuan.cell(row=j,
                                                                                                     column=i).value
                    logging.debug('2卷成绩已搬到tongji表')
                    # 对总分进行统计
                    for kk in range(3, self.shyijuan.max_row + 1):
                        if self.shyijuan.cell(row=kk, column=self.shyijuan.max_column).value != None:
                            # print(self.shyijuan.cell(row=kk + 3, column=2).value,
                            #       int(self.shyijuan.cell(row=kk + 3, column=3).value))
                            renfenzhidian2[self.shyijuan.cell(row=kk, column=2).value] = int(
                                self.shyijuan.cell(row=kk, column=self.shyijuan.max_column).value)
                            # print(kk)

                    logging.debug('总分统计完毕')
                    logging.debug(renfenzhidian2.items())
                    ls1 = list(renfenzhidian2.items())
                    # print(ls[0])
                    # 总分排名
                    ls1.sort(key=lambda x: x[1], reverse=True)
                    logging.debug('总分已排序，最高分是：'.format(ls1[0]))
                    logging.debug("名单为{}".format(ls1[0]))
                    qianshimingdan = ''
                    # 输出前9名
                    for i in range(9):
                        if len(ls1) >= 9:
                            qianshimingdan = qianshimingdan + '{}{} '.format(ls1[i][0], ls1[i][1])
                            self.shtongjish.cell(row=24, column=2).value = qianshimingdan
                    logging.debug("前9名已输出")
                    qianshimingdan = ''
                    # 输出9-18名
                    for i in range(9, 18):
                        if len(ls) >= 18:
                            qianshimingdan = qianshimingdan + '{}{} '.format(ls1[i][0], ls1[i][1])
                            self.shtongjish.cell(row=25, column=2).value = qianshimingdan
                    logging.debug('已输出名次')
        # 如果有表'chengjish'，把I卷II卷成绩汇总在'chengjish'表内
        if 'chengjish' in self.chengjiexcel.sheetnames:
            chengjish = self.chengjiexcel['chengjish']
            for i in range(self.shyijuan.max_row):
                for j in range(self.shyijuan.max_column):
                    chengjish.cell(row=i + 1, column=j + 1).value = self.shyijuan.cell(row=i + 1, column=j + 1).value

        self.shyijuan.delete_cols(self.tishu + 5, self.shyijuan.max_row)

        # 保存导出的文件
        if shijuanlaiyuan != '1':
            self.shijianstr = self.shijianstr + shijuanlaiyuan
        self.chengjiexcel.save('./xls/{}.xlsx'.format(self.shijianstr))  # 保存
        logging.debug('已存')
        self.chengjiexcel.close()
        logging.debug(str(imagePath) + '/xls/{}.xlsx'.format(self.shijianstr))
        # 打开文件
        os.startfile(str(imagePath) + '/xls/{}.xlsx'.format(self.shijianstr))

    # 查询单个学生答题情况
    @pyqtSlot()
    def on_pushButton_chaxunxuesheng_clicked(self):
        logging.debug("进入查询学生功能")
        # 从文本框获取学生姓名
        imgstr = self.lineEdit_chaxunxuesheng.text()
        # 打开excel和工作表
        try:
            self.chengjiexcel = openpyxl.load_workbook(self.openexcelname)
            self.shyijuan = self.chengjiexcel['yijuan']
        except:
            print('查询单个学生成绩时，excel文件损坏！')
        # 遍历学生姓名
        for i in range(3, self.shyijuan.max_row + 1):
            # 如果找到学生姓名
            if imgstr == self.shyijuan.cell(row=i, column=2).value:

                # 学生姓名和试卷地址不为空
                if self.shyijuan.cell(row=i, column=3).value != None and self.shyijuan.cell(row=i,
                                                                                            column=self.shyijuan.max_column).value != None:
                    print('找到{}'.format(imgstr))
                    # 读取试卷地址
                    imgstr = str(
                        imagePath.joinpath('bak').joinpath(
                            self.shyijuan.cell(row=i, column=self.shyijuan.max_column).value))
                    logging.debug("地址为：{}".format(imgstr))
                    # 如果试卷存在
                    if Path(imgstr).exists():
                        print(self.lineEdit_chaxunxuesheng.text(), self.shyijuan.cell(row=i, column=3).value, '分',
                              imgstr.split('\\')[-1])
                        # 读取、调整试卷大小
                        result = QPixmap(imgstr).scaled(self.label_shijuan.width(),
                                                        self.label_shijuan.height())
                        logging.debug("调整大小为laber大小")
                        # 显示在label中
                        self.label_shijuan.setPixmap(result)
                        logging.debug("显示在label中")
                else:
                    print('没有找到{}成绩！'.format(imgstr))
        # 关闭excel文件
        self.chengjiexcel.close()

    # 设置坐标按钮
    @pyqtSlot()
    def on_pushButton_shezhizuobiao_clicked(self):
        os.system('zb.exe')


if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    ui = MainWindow()
    ui.show()
    sys.exit(app.exec_())
