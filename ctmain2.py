#作用：自动形成单个学生错题集ct.xlsx,wuli.docx但不发送
from pathlib import Path
import openpyxl,sys
import cuotishijuan
imagePath = Path(sys.argv[0]).parent
filelist=[]
list1=set()
# 遍历试卷文件形成列表
def bianlipic():
    filelist.clear()
    imagePath = Path(sys.argv[0]).parent
    imagePath2 = imagePath.joinpath('xls')
    for fname in [x for x in imagePath2.iterdir()]:  # 遍历文件夹中每个文件名形成列表
        filelist.append(fname)#文件路径列表

    if len(filelist) != 0:
        return filelist
    else:
        print('xls内无试卷！')
# 导出错题列表到ct.xls
def cuotitaochu(file,ss):
    try:
        chengjiexcel = openpyxl.load_workbook(file)  # 打开excel
        ctxcel = openpyxl.load_workbook('ct.xlsx')  # 打开excel
    except:
        print('excel文件损坏！')

    shijuanlaiyuan = ((str(file).split('\\'))[-1]).split('.')[0]
    # print(shijuanlaiyuan)
    if '2022' in chengjiexcel.sheetnames:
        sh1 = chengjiexcel['2022']  # 打开表单一
    else:
        sh1 = chengjiexcel['yijuan']  # 打开表单一
    sh5 = ctxcel['ct']

#统计选择题
    for i in range(3, sh1.max_row + 1):  # 从第3行开始，到最后一行结束
        if sh1.cell(row=i, column=2).value == ss:#如果找到学生
            for kk in range(sh1.max_column - 4):  # 统计选择题   对每一列即每一题进行统计,sh1.max_column - 4为试题数目
                daan = sh1.cell(row=2, column=kk + 4).value
                tihao=str(sh1.cell(row=1, column=kk + 4).value)
                if (daan is not None) and daan.isalpha() and daan.isupper():
                    rowssh5 = sh5.max_row
                    if sh1.cell(row=i, column=kk + 4).value != daan and sh1.cell(row=i, column=kk + 4).value !=None:#如果㤥题做错了，写入两错题表，kk+1为题号
                        # print(shijuanlaiyuan + '-' + str(kk + 1) + '.jpg')
                        sh5.cell(row=rowssh5 + 1 , column=5).value = shijuanlaiyuan + '-' + tihao + '.jpg'
                        sh5.cell(row=rowssh5 + 1 , column=6).value = shijuanlaiyuan + '-' + tihao + 'a.jpg'
                        sh5.cell(row=rowssh5 + 1, column=7).value = '=VLOOKUP(E{},ly!A:d,2,0)'.format(rowssh5 + 1)
                        sh5.cell(row=rowssh5 + 1, column=8).value = '=VLOOKUP(E{},ly!A:d,3,0)'.format(rowssh5 + 1)
                        sh5.cell(row=rowssh5 + 1, column=9).value = sh1.cell(row=i, column=kk + 4).value
            # 统计二卷
            if 'erjuan' in chengjiexcel.sheetnames:
                erjuan = chengjiexcel['erjuan']  # 打开表单一
                for kk in range(erjuan.max_column - 2):  # 统计选择题   对每一列即每一题进行统计,sh1.max_column - 4为试题数目
                    daan = eval(str(erjuan.cell(row=2, column=kk + 2).value))*0.6
                    tihao=str(erjuan.cell(row=1, column=kk + 2).value)
                    if (daan is not None) and (erjuan.cell(row=i, column=kk + 2).value is not None):
                        rowssh5 = sh5.max_row
                        if int(erjuan.cell(row=i, column=kk + 2).value) <  daan:  # 如果㤥题做错了，写入两错题表，kk+1为题号
                            # print(shijuanlaiyuan + '-' + str(kk + 1) + '.jpg')
                            sh5.cell(row=rowssh5 + 1, column=5).value = shijuanlaiyuan + '-' + tihao + '.jpg'
                            sh5.cell(row=rowssh5 + 1, column=6).value = shijuanlaiyuan + '-' + tihao + 'a.jpg'
                            sh5.cell(row=rowssh5 + 1, column=7).value = '=VLOOKUP(E{},ly!A:d,2,0)'.format(rowssh5 + 1)
                            sh5.cell(row=rowssh5 + 1, column=8).value = '=VLOOKUP(E{},ly!A:d,3,0)'.format(rowssh5 + 1)
                            sh5.cell(row=rowssh5 + 1, column=9).value = erjuan.cell(row=i, column=kk + 2).value

    ctxcel.save('ct.xlsx')  # 保存
    # ctxcel.save('{}.xlsx'.format(ss))
    ctxcel.close()
    chengjiexcel.close()
#清理删除错题列表ct.xls
def qingliexls():
    try:
        chengjiexcel = openpyxl.load_workbook('ct.xlsx')  # 打开excel
    except:
        print('excel文件损坏！')
    sh2 = chengjiexcel['ct']
    sh3 = chengjiexcel['ctxuesheng']  # 打开表单二
    # sh8 = chengjiexcel['ct2']
    print('错题个数为：',sh2.max_row-1)
    # sh8.delete_rows(2, sh8.max_row)
    sh2.delete_rows(2,sh2.max_row)  # 删除行
    sh3.delete_rows(2, sh3.max_row)  # 删除行
    # print(sh2.max_row, sh3.max_row)


    try:
        chengjiexcel.save('ct.xlsx')
        chengjiexcel.close()
    except:
        print('保存出错，先关闭excel再重试！')
def main():
    bianlipic()  # 形成文件列表
    while True:
        ss=input('请输入学生姓名(退出请输入q后回车)：')
        if ss=='q':#输入q则退出
            sys.exit()
        for i in filelist:
            cuotitaochu(i, ss)  # 导出错题列表ct.xls
        cuotishijuan.scdoc()  # 错题导出到word
        qingliexls()  # 清理ct.xls
        list1.clear()
if __name__ == '__main__':
    main()
