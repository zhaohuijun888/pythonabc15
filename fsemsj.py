# 发送试卷。：需要修改自己的邮箱信息
import smtplib
import time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
import sys
import openpyxl

imagePath = Path(sys.argv[0]).parent


def fsem(wjname):
    lctime = time.localtime()

    ymdtime = time.strftime("%Y-%m-%d", lctime)
    wb = openpyxl.load_workbook(imagePath.joinpath(wjname))  # 获取名字，email，二卷
    ws = wb['yijuan']
    emaildz = wb['email']
    rows = ws.max_row
    # print(rows)
    for i in range(3, rows + 1):  # 从第2行开始，到第3行结束
        xsname = ws.cell(row=i, column=2).value
        email = emaildz.cell(row=i, column=2).value
        wjdzhi = ws.cell(row=i, column=ws.max_column).value
        if wjdzhi != None and email != None:
            fujian1 = wjdzhi

            # 1.创建一个带附件的实例
            msg = MIMEMultipart()
            # 2. 加邮件头
            msg['From'] = '18704935399@sohu.com'  # 1修改成自己的邮箱
            msg['To'] = email
            msg['Subject'] = "高二（7）班物理{}试卷".format(ymdtime) + xsname  # 2修改成自己的标题
            # 3.构造附件1（附件为JPG格式的图片）
            att2 = MIMEText(open(fujian1, 'rb').read(), 'base64', 'utf-8')
            att2["Content-Type"] = 'application/octet-stream'
            att2["Content-Disposition"] = 'attachment; filename="wulisj.jpg"'
            msg.attach(att2)
            # 4.发送邮件
            server = smtplib.SMTP_SSL('smtp.sohu.com', 465)  # 3修改为自己的邮箱smtp
            # print('1')
            try:
                # print('2')
                server.login('18704935399@sohu.com', 'BCGX')  # 1修改为自己的邮箱和制授权码
                # print('3')
                server.send_message(msg)  # 发邮箱、接邮箱、纯文本内容
                print('已发送{}'.format(xsname))
                server.quit()
                time.sleep(1)
            except Exception:
                print('发送{}失败！！！！！！！！！！'.format(xsname))
                pass


if __name__ == "__main__":
    fsem()
